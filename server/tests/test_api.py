from __future__ import annotations

from io import BytesIO
from zipfile import ZipFile

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

from app.main import app


pytestmark = pytest.mark.anyio


@pytest.fixture
def anyio_backend() -> str:
    return "asyncio"


@pytest.fixture
async def client() -> AsyncClient:
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://testserver") as api_client:
        yield api_client


async def test_health_returns_ok(client: AsyncClient) -> None:
    response = await client.get("/health")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


async def test_country_packs_include_milestone_1_packs(client: AsyncClient) -> None:
    response = await client.get("/api/country-packs")
    assert response.status_code == 200
    pack_ids = {pack["country_pack_id"] for pack in response.json()["country_packs"]}
    assert {"belgium_peppol", "saudi_zatca", "uk_info"} <= pack_ids
    saudi = next(pack for pack in response.json()["country_packs"] if pack["country_pack_id"] == "saudi_zatca")
    assert "legal_regime_summary" in saudi
    assert "v1_app_capability" in saudi


async def test_template_download_contains_required_sheets(client: AsyncClient) -> None:
    response = await client.get("/api/templates/workbook")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert {"entities", "customers", "invoice_header", "invoice_lines"} <= set(workbook.sheetnames)
    header = [cell.value for cell in workbook["invoice_header"][1]]
    sample = [cell.value for cell in workbook["invoice_header"][2]]
    assert "invoice_time" in header
    assert sample[header.index("selected_country_pack")] == "belgium_peppol"
    assert workbook["invoice_lines"].max_row == 2


async def test_exported_template_supports_belgium_demo_flow(client: AsyncClient) -> None:
    template_response = await client.get("/api/templates/workbook")
    assert template_response.status_code == 200

    upload_response = await client.post(
        "/api/uploads",
        data={"selected_country_pack": "belgium_peppol"},
        files={
            "file": (
                "e-invoicing-v1-template.xlsx",
                template_response.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 200
    payload = upload_response.json()
    assert payload["validation_report"]["summary"]["overall_status"] == "passed"

    generation_response = await client.post(f"/api/uploads/{payload['upload_id']}/generate")
    assert generation_response.status_code == 200
    bundle_response = await client.get(f"/api/uploads/{payload['upload_id']}/evidence-bundle/download")
    assert bundle_response.status_code == 200
    with ZipFile(BytesIO(bundle_response.content)) as archive:
        assert "invoice.xml" in archive.namelist()


async def test_upload_valid_workbook_builds_canonical_json(client: AsyncClient) -> None:
    workbook_bytes = _sample_workbook()
    response = await client.post(
        "/api/uploads",
        data={"selected_country_pack": "saudi_zatca"},
        files={
            "file": (
                "sample.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["canonical_invoice"]["invoice"]["invoice_number"] == "INV-SA-2026-001"
    assert payload["validation_report"]["summary"]["official_artefact_validation"] == "not_configured"
    assert payload["validation_report"]["summary"]["warnings_ack_required"] >= 1
    assert payload["evidence_bundle_preview"]["status"] == "skeleton_only_milestone_1"


async def test_missing_required_sheet_returns_clear_error(client: AsyncClient) -> None:
    workbook = Workbook()
    workbook.active.title = "entities"
    stream = BytesIO()
    workbook.save(stream)
    response = await client.post(
        "/api/uploads",
        data={"selected_country_pack": "belgium_peppol"},
        files={
            "file": (
                "missing-sheets.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    messages = [result["message"] for result in response.json()["validation_report"]["results"]]
    assert "Missing required workbook sheet: customers." in messages


def _sample_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    entities = workbook.create_sheet("entities")
    entities.append(
        [
            "entity_id",
            "legal_name",
            "country_code",
            "tax_registration_number",
            "address_line_1",
            "city",
        ]
    )
    entities.append(["SELLER-1", "Demo Saudi Services LLC", "SA", "300000000000003", "King Road", "Riyadh"])

    customers = workbook.create_sheet("customers")
    customers.append(
        [
            "customer_id",
            "legal_name",
            "buyer_type",
            "country_code",
            "address_line_1",
            "city",
            "tax_registration_number",
        ]
    )
    customers.append(["BUYER-1", "Demo Saudi Buyer LLC", "business", "SA", "Buyer Road", "Riyadh", "300000000000004"])

    header = workbook.create_sheet("invoice_header")
    header.append(
        [
            "invoice_id",
            "invoice_number",
            "invoice_date",
            "invoice_time",
            "entity_id",
            "customer_id",
            "invoice_type",
            "supply_type",
            "transaction_type",
            "selected_country_pack",
            "selected_output_profile",
            "invoice_currency_code",
            "tax_currency_code",
            "net_total",
            "tax_total",
            "gross_total",
        ]
    )
    header.append(
        [
            "INV-1",
            "INV-SA-2026-001",
            "2026-06-24",
            "10:30:00",
            "SELLER-1",
            "BUYER-1",
            "standard_sales_invoice",
            "services",
            "domestic_b2b",
            "saudi_zatca",
            "zatca_standard_tax_invoice_xml_plus_arabic_visual_pdf",
            "SAR",
            "SAR",
            10000,
            1500,
            11500,
        ]
    )

    lines = workbook.create_sheet("invoice_lines")
    lines.append(
        [
            "invoice_id",
            "line_number",
            "description",
            "quantity",
            "unit_code",
            "unit_price",
            "line_net_amount",
            "tax_category_code",
            "tax_rate",
            "tax_amount",
        ]
    )
    lines.append(["INV-1", 1, "Consulting services", 10, "HUR", 1000, 10000, "S", 15, 1500])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
