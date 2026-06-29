from __future__ import annotations

from io import BytesIO
import json
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from httpx import ASGITransport, AsyncClient
import pytest

from app.main import app
from tests.workbook_fixtures import belgium_valid_workbook_bytes


pytestmark = pytest.mark.anyio


@pytest.fixture
def anyio_backend() -> str:
    return "asyncio"


@pytest.fixture
async def client() -> AsyncClient:
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://testserver") as api_client:
        yield api_client


async def test_valid_belgium_sample_workbook_passes_validation(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes())

    assert payload["status"] == "validated"
    assert payload["validation_report"]["summary"]["overall_status"] == "passed"
    assert payload["validation_report"]["summary"]["blocking_errors"] == 0
    assert payload["validation_report"]["summary"]["warnings_ack_required"] == 0


async def test_missing_entities_sheet_is_blocking(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes(remove_sheet="entities"))

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "WB-SHEET-001")
    assert any("Missing required workbook sheet: entities." == result["message"] for result in _results(payload))


async def test_missing_invoice_number_is_blocking(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes(invoice_number=None))

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "GEN-INV-001")


async def test_missing_required_column_is_blocking(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(
        client,
        belgium_valid_workbook_bytes(remove_column=("invoice_lines", "unit_price")),
    )

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "WB-COLUMN-001")
    assert any(
        "Missing required column 'unit_price' on sheet 'invoice_lines'." == result["message"]
        for result in _results(payload)
    )


async def test_duplicate_invoice_number_is_blocking(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes(duplicate_invoice_number=True))

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "GEN-INV-002")


async def test_missing_buyer_reference_and_purchase_order_reference_is_blocking(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(
        client,
        belgium_valid_workbook_bytes(buyer_reference="", purchase_order_reference=""),
    )

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "BE-EINV-011")


async def test_vat_total_mismatch_is_blocking(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes(tax_total=211.00))

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "BE-ARITH-002")


async def test_missing_peppol_ids_requires_acknowledgement_not_blocking(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(
        client,
        belgium_valid_workbook_bytes(seller_peppol_id="", buyer_peppol_id=""),
    )

    assert payload["validation_report"]["summary"]["blocking_errors"] == 0
    assert payload["validation_report"]["summary"]["warnings_ack_required"] == 2
    assert _has_failed_rule(payload, "BE-PEPPOL-001")
    assert _has_failed_rule(payload, "BE-PEPPOL-002")


async def test_canonical_invoice_contains_expected_sections(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes())
    canonical = payload["canonical_invoice"]

    assert canonical["seller"]["legal_name"] == "Demo Belgium Services BV"
    assert canonical["buyer"]["legal_name"] == "Demo Belgium Buyer NV"
    assert canonical["invoice"]["invoice_number"] == "INV-BE-2026-001"
    assert canonical["lines"][0]["description"] == "Consulting services"
    assert canonical["tax_summary"] == [
        {
            "tax_category_code": "S",
            "tax_rate": "21",
            "taxable_amount": "1000.00",
            "tax_amount": "210.00",
        }
    ]
    assert canonical["totals"]["net_total"] == 1000
    assert canonical["totals"]["tax_total"] == 210
    assert canonical["totals"]["gross_total"] == 1210

    canonical_response = await client.get(f"/api/uploads/{payload['upload_id']}/canonical-invoice")
    assert canonical_response.status_code == 200
    assert canonical_response.json()["invoice"]["invoice_number"] == "INV-BE-2026-001"


async def test_evidence_skeleton_includes_persisted_core_artefacts(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes())
    files = {file["filename"]: file for file in payload["evidence_bundle_preview"]["files"]}

    assert files["source_upload_snapshot.xlsx"]["status"] == "stored"
    assert files["source_upload_snapshot.xlsx"]["sha256"]
    assert files["source_upload_snapshot.xlsx"]["storage_path"].endswith(".xlsx")
    assert files["canonical_invoice.json"]["status"] == "stored"
    assert files["canonical_invoice.json"]["sha256"]
    assert files["validation_report.json"]["status"] == "stored"
    assert files["validation_report.json"]["sha256"]
    assert payload["stored_workbook_path"].endswith(".xlsx")
    assert payload["canonical_json_path"].endswith("_canonical_invoice.json")
    assert payload["validation_report_path"].endswith("_validation_report.json")


async def test_evidence_zip_download_contains_skeleton_artefacts(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes())

    response = await client.get(f"/api/uploads/{payload['upload_id']}/evidence-bundle/download")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"
    with ZipFile(BytesIO(response.content)) as archive:
        names = set(archive.namelist())
        assert "source_upload_snapshot.xlsx" in names
        assert "canonical_invoice.json" in names
        assert "validation_report.json" in names
        assert "evidence.json" in names
        assert "country_pack_manifest.json" in names
        assert "hashes.txt" in names
        assert "invoice.xml" not in names


async def test_valid_belgium_sample_generates_xml(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes())

    generate_response = await client.post(f"/api/uploads/{payload['upload_id']}/generate")

    assert generate_response.status_code == 200
    evidence_files = {file["filename"]: file for file in generate_response.json()["files"]}
    assert evidence_files["invoice.xml"]["status"] == "stored"
    assert evidence_files["invoice.xml"]["sha256"]
    assert evidence_files["invoice.xml"]["storage_path"].endswith("_belgium_peppol_invoice.xml")


async def test_generated_xml_contains_expected_belgium_values(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes())
    await client.post(f"/api/uploads/{payload['upload_id']}/generate")

    xml_response = await client.get(f"/api/uploads/{payload['upload_id']}/generated-xml")
    xml = xml_response.text

    assert xml_response.status_code == 200
    ET.fromstring(xml)
    assert "urn:oasis:names:specification:ubl:schema:xsd:Invoice-2" in xml
    assert "<cbc:CustomizationID>urn:cen.eu:en16931:2017#compliant#urn:fdc:peppol.eu:2017:poacc:billing:3.0</cbc:CustomizationID>" in xml
    assert "<cbc:ProfileID>urn:fdc:peppol.eu:2017:poacc:billing:01:1.0</cbc:ProfileID>" in xml
    assert "<cbc:ID>INV-BE-2026-001</cbc:ID>" in xml
    assert "<cbc:CompanyID>BE0123456789</cbc:CompanyID>" in xml
    assert "<cbc:CompanyID>BE0987654321</cbc:CompanyID>" in xml
    assert "<cbc:BuyerReference>BE-BUYER-REF-001</cbc:BuyerReference>" in xml
    assert '<cbc:TaxAmount currencyID="EUR">210.00</cbc:TaxAmount>' in xml
    assert '<cbc:PayableAmount currencyID="EUR">1210.00</cbc:PayableAmount>' in xml


async def test_evidence_zip_includes_generated_belgium_xml_after_generation(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes())
    await client.post(f"/api/uploads/{payload['upload_id']}/generate")

    response = await client.get(f"/api/uploads/{payload['upload_id']}/evidence-bundle/download")

    assert response.status_code == 200
    with ZipFile(BytesIO(response.content)) as archive:
        names = set(archive.namelist())
        assert "invoice.xml" in names
        xml = archive.read("invoice.xml").decode("utf-8")
        assert "INV-BE-2026-001" in xml
        evidence = json.loads(archive.read("evidence.json"))
        assert evidence["selected_country_pack"] == "belgium_peppol"
        assert evidence["country_pack_version"]
        assert evidence["source_workbook"]["filename"] == "BE-VALID-001.xlsx"
        assert evidence["validation"]["internal_validation"] == "passed"
        assert evidence["official_artefact_validation"]["status"] == "not_configured"
        assert "Peppol Schematron validation has run" in evidence["official_artefact_validation"]["note"]
        assert any(output["filename"] == "invoice.xml" for output in evidence["generated_outputs"])


async def test_generation_is_blocked_when_validation_has_blocking_errors(client: AsyncClient) -> None:
    payload = await _upload_belgium_workbook(client, belgium_valid_workbook_bytes(invoice_number=None))

    response = await client.post(f"/api/uploads/{payload['upload_id']}/generate")

    assert response.status_code == 409
    assert "Blocking validation errors" in response.text


async def test_missing_seller_vat_is_blocking(client: AsyncClient) -> None:
    workbook_bytes = _blank_seller_vat_workbook()
    payload = await _upload_belgium_workbook(client, workbook_bytes)

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "GEN-SELLER-002")


async def test_missing_buyer_vat_is_blocking(client: AsyncClient) -> None:
    workbook_bytes = _blank_buyer_vat_workbook()
    payload = await _upload_belgium_workbook(client, workbook_bytes)

    assert payload["validation_report"]["summary"]["blocking_errors"] >= 1
    assert _has_failed_rule(payload, "BE-LEGAL-002")


async def _upload_belgium_workbook(client: AsyncClient, workbook_bytes: bytes) -> dict:
    response = await client.post(
        "/api/uploads",
        data={"selected_country_pack": "belgium_peppol"},
        files={
            "file": (
                "BE-VALID-001.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    return response.json()


def _results(payload: dict) -> list[dict]:
    return payload["validation_report"]["results"]


def _has_failed_rule(payload: dict, rule_id: str) -> bool:
    return any(result["rule_id"] == rule_id and result["status"] == "failed" for result in _results(payload))


def _blank_seller_vat_workbook() -> bytes:
    from openpyxl import load_workbook
    from io import BytesIO

    stream = BytesIO(belgium_valid_workbook_bytes())
    workbook = load_workbook(stream)
    headers = [cell.value for cell in workbook["entities"][1]]
    vat_index = headers.index("tax_registration_number") + 1
    workbook["entities"].cell(row=2, column=vat_index).value = ""
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _blank_buyer_vat_workbook() -> bytes:
    from openpyxl import load_workbook
    from io import BytesIO

    stream = BytesIO(belgium_valid_workbook_bytes())
    workbook = load_workbook(stream)
    headers = [cell.value for cell in workbook["customers"][1]]
    vat_index = headers.index("tax_registration_number") + 1
    workbook["customers"].cell(row=2, column=vat_index).value = ""
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
