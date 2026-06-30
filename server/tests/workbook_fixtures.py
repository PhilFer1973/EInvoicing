from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook


def belgium_valid_workbook_bytes(**overrides: Any) -> bytes:
    workbook = belgium_valid_workbook(**overrides)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def belgium_einvoicebe_validation_workbook_bytes(**overrides: Any) -> bytes:
    workbook = belgium_valid_workbook(
        seller_legal_name="Test Company BV",
        seller_vat="BE0990251719",
        seller_company_number="0990251719",
        seller_enterprise_number="0990251719",
        seller_peppol_id="0208:0990251719",
        **overrides,
    )
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def belgium_einvoicebe_send_workbook_bytes(**overrides: Any) -> bytes:
    defaults = {
        "seller_legal_name": "Test Company BV",
        "seller_vat": "BE0990251719",
        "seller_company_number": "0990251719",
        "seller_enterprise_number": "0990251719",
        "seller_peppol_id": "0208:0990251719",
        "seller_einvoicebe_sender_peppol_id": "0208:099025170",
    }
    defaults.update(overrides)
    workbook = belgium_valid_workbook(**defaults)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def belgium_valid_workbook(
    *,
    remove_sheet: str | None = None,
    invoice_number: str | None = "INV-BE-2026-001",
    buyer_reference: str | None = "BE-BUYER-REF-001",
    purchase_order_reference: str | None = "",
    tax_total: float = 210.00,
    seller_legal_name: str = "Demo Belgium Services BV",
    seller_vat: str = "BE0990251719",
    seller_company_number: str = "0990251719",
    seller_enterprise_number: str = "0990251719",
    seller_peppol_id: str | None = "0208:0990251719",
    seller_einvoicebe_sender_peppol_id: str | None = None,
    buyer_vat: str = "BE0987654394",
    buyer_company_number: str = "0987654394",
    buyer_enterprise_number: str = "0987654394",
    buyer_peppol_id: str | None = "0208:0987654394",
    duplicate_invoice_number: bool = False,
    remove_column: tuple[str, str] | None = None,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    entities = workbook.create_sheet("entities")
    entities.append(
        [
            "entity_id",
            "legal_name",
            "trading_name",
            "country_code",
            "tax_registration_number",
            "company_number",
            "enterprise_number",
            "legal_registration_number",
            "legal_registration_scheme_id",
            "address_line_1",
            "address_line_2",
            "city",
            "postal_code",
            "country_name",
            "email",
            "phone",
            "iban",
            "bic",
            "payment_account_name",
            "peppol_id",
            "peppol_scheme_id",
            "einvoicebe_sender_peppol_id",
        ]
    )
    entities.append(
        [
            "SELLER-BE-1",
            seller_legal_name,
            "",
            "BE",
            seller_vat,
            seller_company_number,
            seller_enterprise_number,
            seller_enterprise_number,
            "0208",
            "Rue Demo 1",
            "",
            "Brussels",
            "1000",
            "Belgium",
            "billing@example.test",
            "",
            "BE68539007547034",
            "BBRUBEBB",
            seller_legal_name,
            seller_peppol_id,
            "0208" if seller_peppol_id else "",
            seller_einvoicebe_sender_peppol_id or "",
        ]
    )

    customers = workbook.create_sheet("customers")
    customers.append(
        [
            "customer_id",
            "legal_name",
            "buyer_type",
            "country_code",
            "tax_registration_number",
            "company_number",
            "enterprise_number",
            "legal_registration_number",
            "legal_registration_scheme_id",
            "address_line_1",
            "address_line_2",
            "city",
            "postal_code",
            "country_name",
            "email",
            "peppol_id",
            "peppol_scheme_id",
        ]
    )
    customers.append(
        [
            "BUYER-BE-1",
            "Demo Belgium Buyer NV",
            "business",
            "BE",
            buyer_vat,
            buyer_company_number,
            buyer_enterprise_number,
            buyer_enterprise_number,
            "0208",
            "Buyer Street 10",
            "",
            "Antwerp",
            "2000",
            "Belgium",
            "ap@example.test",
            buyer_peppol_id,
            "0208" if buyer_peppol_id else "",
        ]
    )

    header = workbook.create_sheet("invoice_header")
    header.append(
        [
            "invoice_id",
            "invoice_number",
            "invoice_date",
            "due_date",
            "tax_point_date",
            "entity_id",
            "customer_id",
            "invoice_type",
            "invoice_type_code",
            "supply_type",
            "transaction_type",
            "selected_country_pack",
            "selected_output_profile",
            "invoice_currency_code",
            "tax_currency_code",
            "net_total",
            "tax_total",
            "gross_total",
            "buyer_reference",
            "purchase_order_reference",
            "payment_means_code",
            "payment_id",
            "payment_terms_note",
            "erp_source_system",
            "erp_document_reference",
            "notes",
        ]
    )
    header_row = [
        "INV-BE-1",
        invoice_number,
        "2026-06-24",
        "2026-07-24",
        "2026-06-24",
        "SELLER-BE-1",
        "BUYER-BE-1",
        "standard_sales_invoice",
        "380",
        "services",
        "domestic_b2b",
        "belgium_peppol",
        "peppol_bis_billing_3_0_ubl_invoice",
        "EUR",
        "EUR",
        1000.00,
        tax_total,
        1210.00,
        buyer_reference,
        purchase_order_reference,
        "30",
        "INV-BE-2026-001",
        "Payment due within 30 days.",
        "Demo ERP",
        "ERP-BE-001",
        "Belgium V1 domestic B2B services invoice.",
    ]
    header.append(header_row)
    if duplicate_invoice_number:
        second_row = header_row.copy()
        second_row[0] = "INV-BE-2"
        header.append(second_row)

    lines = workbook.create_sheet("invoice_lines")
    lines.append(
        [
            "invoice_id",
            "line_number",
            "description",
            "item_name",
            "quantity",
            "unit_code",
            "unit_price",
            "line_net_amount",
            "tax_category_code",
            "tax_rate",
            "tax_amount",
            "line_gross_amount",
        ]
    )
    lines.append(["INV-BE-1", 1, "Consulting services", "Consulting services", 10, "HUR", 100.00, 1000.00, "S", 21, 210.00, 1210.00])

    if remove_sheet:
        del workbook[remove_sheet]
    if remove_column and remove_column[0] in workbook.sheetnames:
        sheet_name, column_name = remove_column
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        if column_name in headers:
            sheet.delete_cols(headers.index(column_name) + 1)

    return workbook


def saudi_valid_workbook_bytes(**overrides: Any) -> bytes:
    workbook = saudi_valid_workbook(**overrides)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def uk_valid_workbook_bytes(**overrides: Any) -> bytes:
    workbook = uk_valid_workbook(**overrides)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def uk_valid_workbook(
    *,
    invoice_number: str | None = "INV-UK-2029-001",
    tax_total: float = 200.00,
    selected_output_profile: str = "storecove_peppol_sandbox_readiness_test",
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    entities = workbook.create_sheet("entities")
    entities.append(
        [
            "entity_id",
            "legal_name",
            "country_code",
            "tax_registration_number",
            "legal_registration_number",
            "address_line_1",
            "city",
            "postal_code",
            "country_name",
            "email",
            "peppol_id",
            "peppol_scheme_id",
        ]
    )
    entities.append(
        [
            "SELLER-UK-1",
            "Demo UK Services Ltd",
            "GB",
            "GB123456789",
            "12345678",
            "1 Demo Street",
            "London",
            "EC1A 1AA",
            "United Kingdom",
            "billing@example.test",
            "0088:123456789",
            "0088",
        ]
    )

    customers = workbook.create_sheet("customers")
    customers.append(
        [
            "customer_id",
            "legal_name",
            "buyer_type",
            "country_code",
            "tax_registration_number",
            "legal_registration_number",
            "address_line_1",
            "city",
            "postal_code",
            "country_name",
            "email",
            "peppol_id",
            "peppol_scheme_id",
        ]
    )
    customers.append(
        [
            "BUYER-UK-1",
            "Demo UK Buyer Ltd",
            "business",
            "GB",
            "GB987654321",
            "87654321",
            "2 Buyer Road",
            "Manchester",
            "M1 1AA",
            "United Kingdom",
            "ap@example.test",
            "0088:987654321",
            "0088",
        ]
    )

    header = workbook.create_sheet("invoice_header")
    header.append(
        [
            "invoice_id",
            "invoice_number",
            "invoice_date",
            "due_date",
            "entity_id",
            "customer_id",
            "invoice_type",
            "invoice_type_code",
            "supply_type",
            "transaction_type",
            "selected_country_pack",
            "selected_output_profile",
            "invoice_currency_code",
            "tax_currency_code",
            "net_total",
            "tax_total",
            "gross_total",
            "buyer_reference",
            "purchase_order_reference",
            "payment_means_code",
            "payment_id",
            "notes",
        ]
    )
    header.append(
        [
            "INV-UK-1",
            invoice_number,
            "2026-06-26",
            "2026-07-26",
            "SELLER-UK-1",
            "BUYER-UK-1",
            "standard_sales_invoice",
            "380",
            "services",
            "domestic_b2b",
            "uk_info",
            selected_output_profile,
            "GBP",
            "GBP",
            1000.00,
            tax_total,
            1200.00,
            "UK-BUYER-REF-001",
            "",
            "30",
            "INV-UK-2029-001",
            "UK Peppol sandbox readiness sample.",
        ]
    )

    lines = workbook.create_sheet("invoice_lines")
    lines.append(
        [
            "invoice_id",
            "line_number",
            "description",
            "item_name",
            "quantity",
            "unit_code",
            "unit_price",
            "line_net_amount",
            "tax_category_code",
            "tax_rate",
            "tax_amount",
            "line_gross_amount",
        ]
    )
    lines.append(["INV-UK-1", 1, "Consulting services", "Consulting services", 10, "HUR", 100.00, 1000.00, "S", 20, 200.00, 1200.00])

    return workbook


def saudi_valid_workbook(
    *,
    invoice_number: str | None = "INV-SA-2026-001",
    invoice_date: str | None = "2026-06-24",
    invoice_time: str | None = "10:30:00",
    seller_vat: str | None = "300000000000003",
    buyer_vat: str | None = "300000000000004",
    buyer_name: str | None = "Demo Saudi Buyer LLC",
    buyer_address_line_1: str | None = "Buyer Road",
    buyer_city: str | None = "Riyadh",
    invoice_type: str = "standard_tax_invoice",
    selected_output_profile: str = "zatca_standard_tax_invoice_xml_plus_arabic_visual_pdf",
    tax_total: float = 1500.00,
    line_description: str | None = "Consulting services",
    line_description_ar: str | None = "خدمات استشارية",
    quantity: float | None = 10,
    unit_code: str | None = "HUR",
    unit_price: float | None = 1000.00,
    line_net_amount: float | None = 10000.00,
    tax_category_code: str | None = "S",
    tax_rate: float | None = 15,
    tax_amount: float | None = 1500.00,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    entities = workbook.create_sheet("entities")
    entities.append(
        [
            "entity_id",
            "legal_name",
            "country_code",
            "tax_registration_number",
            "legal_registration_number",
            "address_line_1",
            "city",
            "postal_code",
            "country_name",
            "email",
            "phone",
        ]
    )
    entities.append(
        [
            "SELLER-SA-1",
            "Demo Saudi Services LLC",
            "SA",
            seller_vat,
            "1010000000",
            "King Road",
            "Riyadh",
            "12211",
            "Saudi Arabia",
            "billing@example.test",
            "",
        ]
    )

    customers = workbook.create_sheet("customers")
    customers.append(
        [
            "customer_id",
            "legal_name",
            "buyer_type",
            "country_code",
            "tax_registration_number",
            "legal_registration_number",
            "address_line_1",
            "city",
            "postal_code",
            "country_name",
            "email",
        ]
    )
    customers.append(
        [
            "BUYER-SA-1",
            buyer_name,
            "business",
            "SA",
            buyer_vat,
            "1010000001",
            buyer_address_line_1,
            buyer_city,
            "12212",
            "Saudi Arabia",
            "ap@example.test",
        ]
    )

    header = workbook.create_sheet("invoice_header")
    header.append(
        [
            "invoice_id",
            "invoice_number",
            "invoice_date",
            "invoice_time",
            "due_date",
            "entity_id",
            "customer_id",
            "invoice_type",
            "invoice_type_code",
            "supply_type",
            "transaction_type",
            "selected_country_pack",
            "selected_output_profile",
            "invoice_currency_code",
            "tax_currency_code",
            "net_total",
            "tax_total",
            "gross_total",
            "payment_means_code",
            "payment_id",
            "notes",
        ]
    )
    header.append(
        [
            "INV-SA-1",
            invoice_number,
            invoice_date,
            invoice_time,
            "2026-07-24",
            "SELLER-SA-1",
            "BUYER-SA-1",
            invoice_type,
            "388" if "simplified" in invoice_type else "380",
            "services",
            "domestic_b2b",
            "saudi_zatca",
            selected_output_profile,
            "SAR",
            "SAR",
            10000.00,
            tax_total,
            11500.00,
            "30",
            "INV-SA-2026-001",
            "Saudi V1 standard B2B tax invoice.",
        ]
    )

    lines = workbook.create_sheet("invoice_lines")
    lines.append(
        [
            "invoice_id",
            "line_number",
            "description",
            "description_ar",
            "item_name",
            "quantity",
            "unit_code",
            "unit_price",
            "line_net_amount",
            "tax_category_code",
            "tax_rate",
            "tax_amount",
            "line_gross_amount",
        ]
    )
    lines.append(
        [
            "INV-SA-1",
            1,
            line_description,
            line_description_ar,
            line_description,
            quantity,
            unit_code,
            unit_price,
            line_net_amount,
            tax_category_code,
            tax_rate,
            tax_amount,
            11500.00,
        ]
    )

    return workbook
