from __future__ import annotations

import hashlib
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from app.adapters.registry import get_adapter
from app.models.canonical import CanonicalInvoice, TaxSummaryLine
from app.models.country_pack import CountryPack
from app.models.upload import UploadRecord
from app.models.validation import ValidationResult
from app.storage.file_store import relative_storage_path, save_binary, save_json
from app.services.validation import boundary_results_for_pack, build_validation_report


REQUIRED_SHEETS = ["entities", "customers", "invoice_header", "invoice_lines"]

REQUIRED_COLUMNS = {
    "entities": [
        "entity_id",
        "legal_name",
        "country_code",
        "tax_registration_number",
        "address_line_1",
        "city",
    ],
    "customers": [
        "customer_id",
        "legal_name",
        "buyer_type",
        "country_code",
        "address_line_1",
        "city",
    ],
    "invoice_header": [
        "invoice_id",
        "invoice_number",
        "invoice_date",
        "entity_id",
        "customer_id",
        "invoice_type",
        "supply_type",
        "transaction_type",
        "selected_country_pack",
        "selected_output_profile",
        "invoice_currency_code",
        "net_total",
        "tax_total",
        "gross_total",
    ],
    "invoice_lines": [
        "invoice_id",
        "line_number",
        "description",
        "quantity",
        "unit_code",
        "unit_price",
        "line_net_amount",
        "tax_category_code",
        "tax_amount",
    ],
}


def parse_workbook_upload(content: bytes, filename: str, pack: CountryPack) -> UploadRecord:
    upload_id = f"UP-{uuid4().hex[:10].upper()}"
    workbook_hash = hashlib.sha256(content).hexdigest()
    results: list[ValidationResult] = []
    canonical_invoice: CanonicalInvoice | None = None
    workbook_path, _ = save_binary("uploads", f"{upload_id}_{_safe_filename(filename)}", content)
    workbook_storage_path = relative_storage_path(workbook_path)

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        results.append(
            ValidationResult(
                rule_id="WB-OPEN-001",
                layer="workbook_structure",
                severity="error",
                status="failed",
                message="The uploaded file could not be opened as an .xlsx workbook.",
                field_path="workbook",
                corrective_action="Upload a valid Excel .xlsx workbook using the V1 template structure.",
                technical_detail=str(exc),
            )
        )
        report = build_validation_report(results)
        evidence = get_adapter(pack.country_pack_id).build_output_placeholder(None)
        evidence = _complete_evidence_preview(evidence, workbook_storage_path, workbook_hash, None, None)
        validation_path, _ = save_json("validation", f"{upload_id}_validation_report.json", report)
        return UploadRecord(
            upload_id=upload_id,
            original_filename=filename,
            selected_country_pack=pack.country_pack_id,
            selected_output_profile=pack.default_output_profile,
            workbook_sha256_hash=workbook_hash,
            status="parse_failed",
            stored_workbook_path=workbook_storage_path,
            validation_report_path=relative_storage_path(validation_path),
            canonical_invoice=None,
            validation_report=report,
            evidence_bundle_preview=evidence,
        )

    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing_sheets:
        for sheet_name in missing_sheets:
            results.append(
                ValidationResult(
                    rule_id="WB-SHEET-001",
                    layer="workbook_structure",
                    severity="error",
                    status="failed",
                    message=f"Missing required workbook sheet: {sheet_name}.",
                    field_path=f"workbook.{sheet_name}",
                    corrective_action="Add the required sheet and upload the workbook again.",
                )
            )
    else:
        results.append(
            ValidationResult(
                rule_id="WB-SHEET-000",
                layer="workbook_structure",
                severity="info",
                status="passed",
                message="Workbook contains the required V1 sheets.",
                field_path="workbook",
            )
        )

    records: dict[str, list[dict[str, Any]]] = {}
    if not missing_sheets:
        for sheet_name in REQUIRED_SHEETS:
            sheet_records, missing_columns = _sheet_records(workbook[sheet_name], sheet_name)
            records[sheet_name] = sheet_records
            for column in missing_columns:
                results.append(
                    ValidationResult(
                        rule_id="WB-COLUMN-001",
                        layer="workbook_structure",
                        severity="error",
                        status="failed",
                        message=f"Missing required column '{column}' on sheet '{sheet_name}'.",
                        field_path=f"{sheet_name}.{column}",
                        corrective_action="Add the required column and upload the workbook again.",
                    )
                )

    if not any(result.severity == "error" and result.status == "failed" for result in results):
        results.extend(_duplicate_invoice_number_results(records, pack))

    if not any(result.severity == "error" and result.status == "failed" for result in results):
        canonical_invoice = _build_canonical_invoice(records, filename, workbook_hash, pack)
        results.append(
            ValidationResult(
                rule_id="CANONICAL-001",
                layer="canonical_construction",
                severity="info",
                status="passed",
                message="Canonical invoice JSON scaffold was constructed from workbook rows.",
                field_path="canonical_invoice",
                country_pack_id=pack.country_pack_id,
                country_pack_version=pack.pack_version,
            )
        )
        mismatch_results = _selected_regime_mismatch_results(canonical_invoice, pack)
        results.extend(mismatch_results)
        if not mismatch_results:
            results.extend(_basic_required_value_results(canonical_invoice, pack))
            if pack.country_pack_id == "belgium_peppol":
                results.extend(_belgium_validation_results(canonical_invoice, records, pack))
            if pack.country_pack_id == "saudi_zatca":
                _apply_saudi_metadata(canonical_invoice)
                results.extend(_saudi_validation_results(canonical_invoice, pack))
            results.extend(boundary_results_for_pack(pack))

    report = build_validation_report(results)
    evidence = get_adapter(pack.country_pack_id).build_output_placeholder(canonical_invoice)
    evidence.generation_id = f"GEN-PREVIEW-{upload_id.replace('UP-', '')}"
    canonical_path = None
    canonical_hash = None
    if canonical_invoice:
        canonical_path, canonical_hash = save_json("canonical", f"{upload_id}_canonical_invoice.json", canonical_invoice)
    validation_path, validation_hash = save_json("validation", f"{upload_id}_validation_report.json", report)
    evidence = _complete_evidence_preview(
        evidence,
        workbook_storage_path,
        workbook_hash,
        (relative_storage_path(canonical_path), canonical_hash) if canonical_path and canonical_hash else None,
        (relative_storage_path(validation_path), validation_hash),
    )
    status = "validated" if report.summary.blocking_errors == 0 else "validation_failed"

    return UploadRecord(
        upload_id=upload_id,
        original_filename=filename,
        selected_country_pack=pack.country_pack_id,
        selected_output_profile=pack.default_output_profile,
        workbook_sha256_hash=workbook_hash,
        status=status,
        stored_workbook_path=workbook_storage_path,
        canonical_json_path=relative_storage_path(canonical_path) if canonical_path else None,
        validation_report_path=relative_storage_path(validation_path),
        canonical_invoice=canonical_invoice,
        validation_report=report,
        evidence_bundle_preview=evidence,
    )


def _sheet_records(sheet: Any, sheet_name: str) -> tuple[list[dict[str, Any]], list[str]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], REQUIRED_COLUMNS[sheet_name]

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_set = set(headers)
    missing_columns = [column for column in REQUIRED_COLUMNS[sheet_name] if column not in header_set]
    records: list[dict[str, Any]] = []

    for row in rows[1:]:
        values = dict(zip(headers, row, strict=False))
        if any(value not in (None, "") for value in values.values()):
            records.append({key: _normalise_cell(value) for key, value in values.items() if key})

    return records, missing_columns


def _build_canonical_invoice(
    records: dict[str, list[dict[str, Any]]],
    filename: str,
    workbook_hash: str,
    pack: CountryPack,
) -> CanonicalInvoice:
    seller = records["entities"][0] if records["entities"] else {}
    buyer = records["customers"][0] if records["customers"] else {}
    invoice = records["invoice_header"][0] if records["invoice_header"] else {}
    invoice_id = invoice.get("invoice_id")
    lines = [
        line
        for line in records["invoice_lines"]
        if not invoice_id or line.get("invoice_id") == invoice_id
    ]

    return CanonicalInvoice(
        invoice=invoice,
        seller=seller,
        buyer=buyer,
        lines=lines,
        tax_summary=_derive_tax_summary(lines),
        totals={
            "net_total": invoice.get("net_total"),
            "tax_total": invoice.get("tax_total"),
            "gross_total": invoice.get("gross_total"),
            "line_extension_total": invoice.get("line_extension_total") or invoice.get("net_total"),
            "tax_exclusive_total": invoice.get("tax_exclusive_total") or invoice.get("net_total"),
            "tax_inclusive_total": invoice.get("tax_inclusive_total") or invoice.get("gross_total"),
            "payable_amount": invoice.get("payable_amount") or invoice.get("gross_total"),
        },
        source={
            "original_filename": filename,
            "workbook_sha256_hash": workbook_hash,
        },
        metadata={
            "country_pack_id": pack.country_pack_id,
            "country_pack_version": pack.pack_version,
            "support_level": pack.support_level,
            "selected_output_profile": invoice.get("selected_output_profile") or pack.default_output_profile,
            "official_artefact_validation": "not_configured",
        },
    )


def _derive_tax_summary(lines: list[dict[str, Any]]) -> list[TaxSummaryLine]:
    grouped: dict[tuple[str, str], dict[str, Decimal]] = defaultdict(lambda: {"base": Decimal("0"), "tax": Decimal("0")})
    for line in lines:
        category = str(line.get("tax_category_code") or "").strip()
        rate = str(line.get("tax_rate") or "0").strip()
        key = (category, rate)
        grouped[key]["base"] += _decimal(line.get("line_net_amount"))
        grouped[key]["tax"] += _decimal(line.get("tax_amount"))

    return [
        TaxSummaryLine(
            tax_category_code=category,
            tax_rate=rate,
            taxable_amount=f"{amounts['base']:.2f}",
            tax_amount=f"{amounts['tax']:.2f}",
        )
        for (category, rate), amounts in sorted(grouped.items())
        if category
    ]


def _duplicate_invoice_number_results(records: dict[str, list[dict[str, Any]]], pack: CountryPack) -> list[ValidationResult]:
    invoice_numbers = [
        str(row.get("invoice_number") or "").strip()
        for row in records.get("invoice_header", [])
        if str(row.get("invoice_number") or "").strip()
    ]
    duplicates = sorted({number for number in invoice_numbers if invoice_numbers.count(number) > 1})
    return [
        ValidationResult(
            rule_id="GEN-INV-002",
            layer="referential_integrity",
            severity="error",
            status="failed",
            message=f"Duplicate invoice number within upload: {number}.",
            field_path="invoice_header.invoice_number",
            country_pack_id=pack.country_pack_id,
            country_pack_version=pack.pack_version,
            corrective_action="Use one unique invoice number per V1 workbook upload.",
        )
        for number in duplicates
    ]


def _selected_regime_mismatch_results(canonical: CanonicalInvoice, pack: CountryPack) -> list[ValidationResult]:
    workbook_pack_id = str(canonical.invoice.get("selected_country_pack") or "").strip()
    workbook_profile = str(canonical.invoice.get("selected_output_profile") or "").strip()
    pack_mismatch = bool(workbook_pack_id and workbook_pack_id != pack.country_pack_id)
    profile_pack_id = _pack_id_for_profile(workbook_profile)
    profile_mismatch = bool(profile_pack_id and profile_pack_id != pack.country_pack_id)

    if not pack_mismatch and not profile_mismatch:
        return []

    selected_label = _regime_label(pack.country_pack_id)
    workbook_label = _regime_label(workbook_pack_id)
    message = "Wrong regime selected" if workbook_label and selected_label else "Workbook does not match regime"
    corrective_action = (
        f"This workbook is for {workbook_label}. Switch to {workbook_label} or upload a {selected_label} workbook."
        if workbook_label and selected_label and pack_mismatch
        else "Switch regime or upload the correct workbook."
    )

    return [
        ValidationResult(
            rule_id="WB-REGIME-001",
            layer="workbook_structure",
            severity="error",
            status="failed",
            message=message,
            field_path="invoice_header.selected_country_pack",
            country_pack_id=pack.country_pack_id,
            country_pack_version=pack.pack_version,
            corrective_action=corrective_action,
            technical_detail=(
                f"Selected country pack={pack.country_pack_id}; workbook country pack={workbook_pack_id or 'not provided'}; "
                f"selected output profile={pack.default_output_profile or 'not configured'}; workbook output profile={workbook_profile or 'not provided'}."
            ),
        )
    ]


def _belgium_validation_results(
    canonical: CanonicalInvoice,
    records: dict[str, list[dict[str, Any]]],
    pack: CountryPack,
) -> list[ValidationResult]:
    results: list[ValidationResult] = []
    seller_vat_missing = _is_blank(canonical.seller.get("tax_registration_number"))
    buyer_vat_missing = _is_blank(canonical.buyer.get("tax_registration_number"))
    buyer_reference_missing = _is_blank(canonical.invoice.get("buyer_reference"))
    po_reference_missing = _is_blank(canonical.invoice.get("purchase_order_reference"))

    if seller_vat_missing:
        results.append(
            _blocking_result(
                "BE-LEGAL-001",
                "legal_invoice_requirements",
                "seller.tax_registration_number",
                "Seller VAT number is required for the Belgian V1 B2B scenario.",
                pack,
                "Add tax_registration_number on the entities sheet.",
            )
        )
    if buyer_vat_missing and str(canonical.buyer.get("buyer_type") or "").lower() in {"business", "government", ""}:
        results.append(
            _blocking_result(
                "BE-LEGAL-002",
                "legal_invoice_requirements",
                "buyer.tax_registration_number",
                "Buyer VAT number is required for Belgian B2B invoices.",
                pack,
                "Add tax_registration_number on the customers sheet.",
            )
        )
    if buyer_reference_missing and po_reference_missing:
        results.append(
            _blocking_result(
                "BE-EINV-011",
                "country_preflight",
                "invoice.buyer_reference",
                "Either buyer reference or purchase order reference must be provided.",
                pack,
                "Add buyer_reference or purchase_order_reference on the invoice_header sheet.",
            )
        )

    results.extend(_belgium_vat_reconciliation_results(canonical, records, pack))
    results.extend(_belgium_peppol_warning_results(canonical, pack))

    if not any(result.severity == "error" and result.status == "failed" for result in results):
        results.append(
            ValidationResult(
                rule_id="BE-PREFLIGHT-000",
                layer="country_preflight",
                severity="info",
                status="passed",
                message="Belgium workbook preflight checks passed.",
                field_path="invoice.selected_country_pack",
                country_pack_id=pack.country_pack_id,
                country_pack_version=pack.pack_version,
            )
        )

    return results


def _belgium_vat_reconciliation_results(
    canonical: CanonicalInvoice,
    records: dict[str, list[dict[str, Any]]],
    pack: CountryPack,
) -> list[ValidationResult]:
    line_net_total = sum(_decimal(line.get("line_net_amount")) for line in canonical.lines)
    line_tax_total = sum(_decimal(line.get("tax_amount")) for line in canonical.lines)
    header_net = _decimal(canonical.invoice.get("net_total"))
    header_tax = _decimal(canonical.invoice.get("tax_total"))
    header_gross = _decimal(canonical.invoice.get("gross_total"))
    results: list[ValidationResult] = []

    if not _money_equal(line_net_total, header_net):
        results.append(
            _blocking_result(
                "BE-ARITH-001",
                "arithmetic",
                "invoice.net_total",
                "Invoice net total does not match the sum of line net amounts.",
                pack,
                "Correct net_total or the line_net_amount values in the workbook.",
            )
        )
    if not _money_equal(line_tax_total, header_tax):
        results.append(
            _blocking_result(
                "BE-ARITH-002",
                "arithmetic",
                "invoice.tax_total",
                "Invoice VAT total does not match the sum of line VAT amounts.",
                pack,
                "Correct tax_total or the line tax_amount values in the workbook.",
            )
        )
    if not _money_equal(header_net + header_tax, header_gross):
        results.append(
            _blocking_result(
                "BE-ARITH-003",
                "arithmetic",
                "invoice.gross_total",
                "Invoice gross total must equal net total plus VAT total.",
                pack,
                "Correct gross_total, net_total or tax_total in the invoice_header sheet.",
            )
        )

    for summary in canonical.tax_summary:
        expected_tax = (_decimal(summary.taxable_amount) * _decimal(summary.tax_rate) / Decimal("100")).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        actual_tax = _decimal(summary.tax_amount)
        if not _money_equal(expected_tax, actual_tax):
            results.append(
                _blocking_result(
                    "BE-ARITH-004",
                    "arithmetic",
                    "tax_summary.tax_amount",
                    f"VAT summary mismatch for category {summary.tax_category_code} at {summary.tax_rate}%.",
                    pack,
                    "Correct the line VAT amount or VAT rate so the VAT category total reconciles.",
                )
            )

    return results


def _belgium_peppol_warning_results(canonical: CanonicalInvoice, pack: CountryPack) -> list[ValidationResult]:
    warnings: list[ValidationResult] = []
    if _is_blank(canonical.seller.get("peppol_id")):
        warnings.append(
            _ack_warning_result(
                "BE-PEPPOL-001",
                "seller.peppol_id",
                "Seller Peppol endpoint ID is missing. This would prevent live Peppol delivery.",
                pack,
            )
        )
    if _is_blank(canonical.buyer.get("peppol_id")):
        warnings.append(
            _ack_warning_result(
                "BE-PEPPOL-002",
                "buyer.peppol_id",
                "Buyer Peppol endpoint ID is missing. This would prevent live Peppol delivery.",
                pack,
            )
        )
    return warnings


def _apply_saudi_metadata(canonical: CanonicalInvoice) -> None:
    canonical.metadata["rounding_policy"] = {
        "mode": "half_up",
        "amount_decimals": 2,
        "vat_summary_validation": "document_level_by_vat_category_and_rate",
    }


def _saudi_validation_results(canonical: CanonicalInvoice, pack: CountryPack) -> list[ValidationResult]:
    results: list[ValidationResult] = []
    invoice = canonical.invoice
    seller = canonical.seller
    buyer = canonical.buyer
    invoice_type = str(invoice.get("invoice_type") or "").lower()
    selected_profile = str(invoice.get("selected_output_profile") or "").lower()

    seller_tin = seller.get("tax_registration_number")
    if _is_blank(seller_tin):
        results.append(
            _blocking_result(
                "SA-SELLER-001",
                "legal_invoice_requirements",
                "seller.tax_registration_number",
                "Seller VAT/TIN is required for Saudi V1 standard B2B invoices.",
                pack,
                "Add tax_registration_number on the entities sheet.",
            )
        )
    elif not _valid_saudi_tin(seller_tin):
        results.append(
            _blocking_result(
                "SA-SELLER-002",
                "legal_invoice_requirements",
                "seller.tax_registration_number",
                "Seller VAT/TIN must be 15 digits and start and end with 3 for the V1 sample rule.",
                pack,
                "Use a 15-digit Saudi VAT/TIN such as 300000000000003.",
            )
        )

    if _is_blank(buyer.get("tax_registration_number")) and str(buyer.get("buyer_type") or "").lower() in {"business", "government", ""}:
        results.append(
            _blocking_result(
                "SA-BUYER-001",
                "legal_invoice_requirements",
                "buyer.tax_registration_number",
                "Buyer VAT/TIN is required for the Saudi V1 B2B standard invoice scenario.",
                pack,
                "Add tax_registration_number on the customers sheet.",
            )
        )
    if _is_blank(buyer.get("legal_name")):
        results.append(
            _blocking_result(
                "SA-BUYER-002",
                "legal_invoice_requirements",
                "buyer.legal_name",
                "Buyer name is required for Saudi V1 standard B2B invoices.",
                pack,
                "Add legal_name on the customers sheet.",
            )
        )
    if _is_blank(buyer.get("address_line_1")) or _is_blank(buyer.get("city")):
        results.append(
            _blocking_result(
                "SA-BUYER-003",
                "legal_invoice_requirements",
                "buyer.address",
                "Buyer address line 1 and city are required for Saudi V1 standard B2B invoices.",
                pack,
                "Add address_line_1 and city on the customers sheet.",
            )
        )

    unsupported_type_tokens = {"simplified", "credit", "debit", "prepayment"}
    if any(token in invoice_type or token in selected_profile for token in unsupported_type_tokens):
        results.append(
            _blocking_result(
                "SA-INV-TYPE-001",
                "country_preflight",
                "invoice.invoice_type",
                "Saudi V1 supports only the standard B2B tax invoice scenario; simplified invoices, credit notes, debit notes and prepayment invoices are blocked.",
                pack,
                "Use the standard tax invoice profile for the V1 Saudi scenario.",
            )
        )

    if not canonical.lines:
        results.append(
            _blocking_result(
                "SA-LINE-000",
                "legal_invoice_requirements",
                "lines",
                "At least one invoice line is required for Saudi V1 validation.",
                pack,
                "Add invoice line rows to the invoice_lines sheet.",
            )
        )
    for index, line in enumerate(canonical.lines, start=1):
        line_path = f"lines[{index}]"
        line_required_checks = [
            ("SA-LINE-001", "description", "Line description is required."),
            ("SA-LINE-002", "quantity", "Line quantity is required."),
            ("SA-LINE-003", "unit_code", "Line unit code is required."),
            ("SA-LINE-004", "unit_price", "Line unit price is required."),
            ("SA-LINE-005", "line_net_amount", "Line net amount is required."),
            ("SA-LINE-006", "tax_category_code", "Line VAT category is required."),
            ("SA-LINE-007", "tax_rate", "Line VAT rate is required."),
            ("SA-LINE-008", "tax_amount", "Line VAT amount is required."),
        ]
        for rule_id, field, message in line_required_checks:
            if _is_blank(line.get(field)):
                results.append(
                    _blocking_result(
                        rule_id,
                        "legal_invoice_requirements",
                        f"{line_path}.{field}",
                        message,
                        pack,
                        f"Add {field} on the invoice_lines sheet.",
                    )
                )

    results.extend(_saudi_vat_reconciliation_results(canonical, pack))

    if not any(result.severity == "error" and result.status == "failed" for result in results):
        results.append(
            ValidationResult(
                rule_id="SA-PREFLIGHT-000",
                layer="country_preflight",
                severity="info",
                status="passed",
                message="Saudi workbook validation foundation checks passed for the V1 standard B2B scenario.",
                field_path="invoice.selected_country_pack",
                country_pack_id=pack.country_pack_id,
                country_pack_version=pack.pack_version,
            )
        )

    return results


def _saudi_vat_reconciliation_results(canonical: CanonicalInvoice, pack: CountryPack) -> list[ValidationResult]:
    line_net_total = sum(_decimal(line.get("line_net_amount")) for line in canonical.lines)
    line_tax_total = sum(_decimal(line.get("tax_amount")) for line in canonical.lines)
    header_net = _decimal(canonical.invoice.get("net_total"))
    header_tax = _decimal(canonical.invoice.get("tax_total"))
    header_gross = _decimal(canonical.invoice.get("gross_total"))
    results: list[ValidationResult] = []

    if not _money_equal(line_net_total, header_net):
        results.append(
            _blocking_result(
                "SA-ARITH-001",
                "arithmetic",
                "invoice.net_total",
                "Saudi invoice net total does not match the sum of line net amounts.",
                pack,
                "Correct net_total or line_net_amount values in the workbook.",
            )
        )
    if not _money_equal(line_tax_total, header_tax):
        results.append(
            _blocking_result(
                "SA-ARITH-002",
                "arithmetic",
                "invoice.tax_total",
                "Saudi invoice VAT total does not match the sum of line VAT amounts.",
                pack,
                "Correct tax_total or line tax_amount values in the workbook.",
            )
        )
    if not _money_equal(header_net + header_tax, header_gross):
        results.append(
            _blocking_result(
                "SA-ARITH-003",
                "arithmetic",
                "invoice.gross_total",
                "Saudi invoice gross total must equal net total plus VAT total.",
                pack,
                "Correct gross_total, net_total or tax_total in the invoice_header sheet.",
            )
        )

    for summary in canonical.tax_summary:
        expected_tax = (_decimal(summary.taxable_amount) * _decimal(summary.tax_rate) / Decimal("100")).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        if not _money_equal(expected_tax, _decimal(summary.tax_amount)):
            results.append(
                _blocking_result(
                    "SA-ARITH-004",
                    "arithmetic",
                    "tax_summary.tax_amount",
                    f"Saudi VAT summary mismatch for category {summary.tax_category_code} at {summary.tax_rate}%.",
                    pack,
                    "Correct the line VAT amount or VAT rate so the VAT category total reconciles.",
                )
            )

    return results


def _blocking_result(
    rule_id: str,
    layer: str,
    field_path: str,
    message: str,
    pack: CountryPack,
    corrective_action: str,
) -> ValidationResult:
    return ValidationResult(
        rule_id=rule_id,
        layer=layer,
        severity="error",
        status="failed",
        message=message,
        field_path=field_path,
        country_pack_id=pack.country_pack_id,
        country_pack_version=pack.pack_version,
        corrective_action=corrective_action,
    )


def _ack_warning_result(rule_id: str, field_path: str, message: str, pack: CountryPack) -> ValidationResult:
    return ValidationResult(
        rule_id=rule_id,
        layer="country_preflight",
        severity="warning_ack_required",
        status="failed",
        message=message,
        field_path=field_path,
        country_pack_id=pack.country_pack_id,
        country_pack_version=pack.pack_version,
        corrective_action="Add Peppol endpoint details or acknowledge before export in a later milestone.",
    )


def _basic_required_value_results(canonical: CanonicalInvoice, pack: CountryPack) -> list[ValidationResult]:
    checks = [
        ("GEN-INV-001", "invoice.invoice_number", canonical.invoice.get("invoice_number"), "Invoice number is required."),
        ("GEN-SELLER-001", "seller.legal_name", canonical.seller.get("legal_name"), "Seller legal name is required."),
        (
            "GEN-SELLER-002",
            "seller.tax_registration_number",
            canonical.seller.get("tax_registration_number"),
            "Seller tax registration number is required.",
        ),
        ("GEN-BUYER-001", "buyer.legal_name", canonical.buyer.get("legal_name"), "Buyer legal name is required."),
        ("GEN-DATE-001", "invoice.invoice_date", canonical.invoice.get("invoice_date"), "Invoice date is required."),
    ]
    results: list[ValidationResult] = []
    for rule_id, field_path, value, message in checks:
        failed = value in (None, "")
        results.append(
            ValidationResult(
                rule_id=rule_id,
                layer="legal_invoice_requirements",
                severity="error" if failed else "info",
                status="failed" if failed else "passed",
                message=message if failed else f"{field_path} is present.",
                field_path=field_path,
                country_pack_id=pack.country_pack_id,
                country_pack_version=pack.pack_version,
                corrective_action="Correct the Excel workbook and upload it again." if failed else None,
            )
        )

    if pack.country_pack_id == "saudi_zatca":
        failed = canonical.invoice.get("invoice_time") in (None, "")
        results.append(
            ValidationResult(
                rule_id="SA-INV-001",
                layer="country_preflight",
                severity="error" if failed else "info",
                status="failed" if failed else "passed",
                message="Saudi V1 requires invoice issue time." if failed else "Saudi invoice issue time is present.",
                field_path="invoice.invoice_time",
                country_pack_id=pack.country_pack_id,
                country_pack_version=pack.pack_version,
                corrective_action="Add invoice_time to the invoice_header sheet." if failed else None,
            )
        )

    return results


def _normalise_cell(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _complete_evidence_preview(
    evidence: Any,
    workbook_storage_path: str,
    workbook_hash: str,
    canonical: tuple[str, str] | None,
    validation: tuple[str, str] | None,
) -> Any:
    for file in evidence.files:
        if file.filename == "source_upload_snapshot.xlsx":
            file.status = "stored"
            file.sha256 = workbook_hash
            file.storage_path = workbook_storage_path
        elif file.filename == "canonical_invoice.json" and canonical:
            file.status = "stored"
            file.sha256 = canonical[1]
            file.storage_path = canonical[0]
        elif file.filename == "validation_report.json" and validation:
            file.status = "stored"
            file.sha256 = validation[1]
            file.storage_path = validation[0]
    return evidence


def _safe_filename(filename: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in {".", "-", "_"} else "_" for char in filename)
    return cleaned or "upload.xlsx"


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _valid_saudi_tin(value: Any) -> bool:
    tin = str(value).strip()
    return len(tin) == 15 and tin.isdigit() and tin.startswith("3") and tin.endswith("3")


def _regime_label(pack_id: str) -> str | None:
    return {
        "belgium_peppol": "Belgium",
        "saudi_zatca": "Saudi",
        "uk_info": "UK",
    }.get(pack_id)


def _pack_id_for_profile(profile_id: str) -> str | None:
    if profile_id.startswith("peppol_"):
        return "belgium_peppol"
    if profile_id.startswith("zatca_"):
        return "saudi_zatca"
    return None


def _money_equal(left: Decimal, right: Decimal) -> bool:
    return left.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) == right.quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )


def _decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")
