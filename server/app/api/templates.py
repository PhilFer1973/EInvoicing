from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from app.services.workbook import REQUIRED_COLUMNS


router = APIRouter(prefix="/api/templates", tags=["templates"])


# A valid Belgium domestic B2B starter row makes the template immediately testable.
# Optional fields are included so the same four-sheet structure can be adapted for Saudi.
SAMPLE_ROWS = {
    "entities": {
        "entity_id": "SELLER-BE-1",
        "legal_name": "Demo Belgium Services BV",
        "country_code": "BE",
        "tax_registration_number": "BE0990251719",
        "company_number": "0990251719",
        "enterprise_number": "0990251719",
        "legal_registration_number": "0990251719",
        "legal_registration_scheme_id": "0208",
        "address_line_1": "Rue Demo 1",
        "city": "Brussels",
        "postal_code": "1000",
        "country_name": "Belgium",
        "peppol_id": "0208:0990251719",
        "peppol_scheme_id": "0208",
    },
    "customers": {
        "customer_id": "BUYER-BE-1",
        "legal_name": "Demo Belgium Buyer NV",
        "buyer_type": "business",
        "country_code": "BE",
        "tax_registration_number": "BE0987654394",
        "company_number": "0987654394",
        "enterprise_number": "0987654394",
        "legal_registration_number": "0987654394",
        "legal_registration_scheme_id": "0208",
        "address_line_1": "Buyer Street 10",
        "city": "Antwerp",
        "postal_code": "2000",
        "country_name": "Belgium",
        "peppol_id": "0208:0987654394",
        "peppol_scheme_id": "0208",
    },
    "invoice_header": {
        "invoice_id": "INV-BE-1",
        "invoice_number": "INV-BE-2026-001",
        "invoice_date": "2026-06-24",
        "invoice_time": "",
        "due_date": "2026-07-24",
        "entity_id": "SELLER-BE-1",
        "customer_id": "BUYER-BE-1",
        "invoice_type": "standard_sales_invoice",
        "invoice_type_code": "380",
        "supply_type": "services",
        "transaction_type": "domestic_b2b",
        "selected_country_pack": "belgium_peppol",
        "selected_output_profile": "peppol_bis_billing_3_0_ubl_invoice",
        "invoice_currency_code": "EUR",
        "tax_currency_code": "EUR",
        "net_total": 1000.00,
        "tax_total": 210.00,
        "gross_total": 1210.00,
        "buyer_reference": "BE-BUYER-REF-001",
        "purchase_order_reference": "",
        "payment_means_code": "30",
        "payment_id": "INV-BE-2026-001",
        "invoice_counter_value": "",
        "previous_invoice_hash": "",
    },
    "invoice_lines": {
        "invoice_id": "INV-BE-1",
        "line_number": 1,
        "description": "Consulting services",
        "description_ar": "",
        "item_name": "Consulting services",
        "quantity": 10,
        "unit_code": "HUR",
        "unit_price": 100.00,
        "line_net_amount": 1000.00,
        "tax_category_code": "S",
        "tax_rate": 21,
        "tax_amount": 210.00,
        "line_gross_amount": 1210.00,
    },
}


@router.get("/workbook")
def download_workbook_template() -> StreamingResponse:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, required_columns in REQUIRED_COLUMNS.items():
        columns = list(dict.fromkeys([*required_columns, *SAMPLE_ROWS[sheet_name]]))
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(columns)
        sheet.append([SAMPLE_ROWS[sheet_name].get(column_name, "") for column_name in columns])
        for column_index, column_name in enumerate(columns, start=1):
            sheet.cell(row=1, column=column_index).style = "Headline 3"
            sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = max(
                16,
                min(34, len(column_name) + 4),
            )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="e-invoicing-v1-template.xlsx"'},
    )
